#!/usr/bin/env python3
"""Build earnings-data.xlsx from build/stocks.json + build/live.json.

Four sheets: All Holdings (master table), Earnings Schedule (chronological),
Calendar (month grids Jul-Dec 2026), About. Run tools/dump-json.mjs first.
"""
import json, os, datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
stocks_doc = json.load(open(os.path.join(ROOT, "build/stocks.json")))
live_doc = json.load(open(os.path.join(ROOT, "build/live.json")))

STOCKS = stocks_doc["stocks"]
MODELS = stocks_doc["models"]          # [{key,label}]
LIVE = live_doc.get("data", {})
REFRESHED = live_doc.get("refreshedAt")

# ---- palette / fonts --------------------------------------------------------
FONT = "Times New Roman"
INK = "1a1a1a"
NAVY = "1f3a5f"
GOLD = "8a7440"
GREEN = "1a7a3a"
RED = "b23a2f"
GREY = "8a8a8a"
HDR_FILL = PatternFill("solid", fgColor="1f3a5f")
BAND_FILL = PatternFill("solid", fgColor="f4f1ea")
EST_FILL = PatternFill("solid", fgColor="f0f0f0")
thin = Side(style="thin", color="d9d4c7")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def base_font(**kw):
    kw.setdefault("name", FONT)
    return Font(**kw)


def hdr_cell(c, text):
    c.value = text
    c.font = base_font(bold=True, color="ffffff", size=11)
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER


def rating(r):
    """r is [label, target] or missing."""
    if not r:
        return "", None
    label = r[0] or ""
    target = r[1] if len(r) > 1 else None
    return label, target


def live_of(t):
    return LIVE.get(t, {})


def next_earnings(t):
    es = live_of(t).get("earnings") or []
    return es[0] if es else None


def fmt_dates(es):
    return "  ·  ".join(
        dt.date.fromisoformat(e["date"]).strftime("%b %d") + (" (est)" if e.get("estimated") else "")
        for e in es
    )


# stocks to include: real equities only (skip funds / delisted / placeholder)
rows = [s for s in STOCKS if s.get("sector") and not s.get("fund") and not s.get("delisted")]
rows.sort(key=lambda s: (s["sector"], s["t"]))

wb = Workbook()

# ============================================================
# Sheet 1 — All Holdings (master table)
# ============================================================
ws = wb.active
ws.title = "All Holdings"

model_keys = [m["key"] for m in MODELS]
model_labels = {m["key"]: m["label"] for m in MODELS}

headers = (
    ["Ticker", "Company", "Sector", "Price", "30D %", "60D %",
     "Next Earnings", "Status", "Source", "All 2026 Dates",
     "WFS Rating", "WFS Tgt", "WFS Upside",
     "ISI Rating", "ISI Tgt", "ISI Upside",
     "M*", "M* Tgt"]
    + [model_labels[k] for k in model_keys]
    + ["Verify?"]
)

for j, h in enumerate(headers, 1):
    hdr_cell(ws.cell(row=1, column=j), h)

r = 2
for s in rows:
    t = s["t"]
    l = live_of(t)
    price = l.get("price")
    ne = next_earnings(t)
    es = l.get("earnings") or []
    wfs_l, wfs_t = rating(s.get("wfs"))
    isi_l, isi_t = rating(s.get("isi"))
    ms = s.get("ms") or []
    ms_stars = ms[0] if ms else None
    ms_tgt = ms[1] if len(ms) > 1 else None

    def upside(tgt):
        if tgt and price:
            return (tgt - price) / price
        return None

    vals = [
        t, s.get("name", ""), s["sector"], price,
        (l.get("perf30") / 100 if l.get("perf30") is not None else None),
        (l.get("perf60") / 100 if l.get("perf60") is not None else None),
        (dt.date.fromisoformat(ne["date"]) if ne else None),
        ("Estimated" if ne and ne.get("estimated") else ("Confirmed" if ne else "TBD")),
        (ne.get("source", "") if ne else ""),
        fmt_dates(es),
        wfs_l, wfs_t, upside(wfs_t),
        isi_l, isi_t, upside(isi_t),
        (("★" * int(ms_stars)) if ms_stars else ""), ms_tgt,
    ] + [("X" if k in (s.get("models") or []) else "") for k in model_keys] + [
        ("⚠" if s.get("verify") else "")
    ]

    band = BAND_FILL if (r % 2 == 0) else None
    for j, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=j, value=v)
        c.font = base_font(size=10, bold=(j == 1))
        c.border = BORDER
        if band:
            c.fill = band
    r += 1

nrows = r - 1
# column formats
col = {h: i + 1 for i, h in enumerate(headers)}
def colL(name): return get_column_letter(col[name])

for rr in range(2, r):
    ws.cell(rr, col["Price"]).number_format = '$#,##0.00'
    ws.cell(rr, col["WFS Tgt"]).number_format = '$#,##0'
    ws.cell(rr, col["ISI Tgt"]).number_format = '$#,##0'
    ws.cell(rr, col["M* Tgt"]).number_format = '$#,##0'
    for pct in ("30D %", "60D %", "WFS Upside", "ISI Upside"):
        cc = ws.cell(rr, col[pct]); cc.number_format = '0.0%;[Red](0.0%)'
        if cc.value is not None:
            cc.font = base_font(size=10, color=(GREEN if cc.value >= 0 else RED))
    ws.cell(rr, col["Next Earnings"]).number_format = 'mmm dd, yyyy'
    ws.cell(rr, col["Next Earnings"]).alignment = Alignment(horizontal="center")
    st = ws.cell(rr, col["Status"])
    st.alignment = Alignment(horizontal="center")
    if st.value == "Estimated":
        st.font = base_font(size=9, italic=True, color=GREY)
    elif st.value == "Confirmed":
        st.font = base_font(size=9, bold=True, color=GREEN)
    for k in model_keys:
        mc = ws.cell(rr, col[model_labels[k]]); mc.alignment = Alignment(horizontal="center")
        mc.font = base_font(size=10, bold=True, color=NAVY)
    ws.cell(rr, col["M*"]).font = base_font(size=10, color=GOLD)
    ws.cell(rr, col["Verify?"]).alignment = Alignment(horizontal="center")

widths = {"Ticker": 9, "Company": 24, "Sector": 22, "Price": 11, "30D %": 8, "60D %": 8,
          "Next Earnings": 15, "Status": 11, "Source": 14, "All 2026 Dates": 30,
          "WFS Rating": 13, "WFS Tgt": 9, "WFS Upside": 10,
          "ISI Rating": 12, "ISI Tgt": 9, "ISI Upside": 10, "M*": 7, "M* Tgt": 9, "Verify?": 8}
for h in headers:
    ws.column_dimensions[colL(h)].width = widths.get(h, 11)
ws.row_dimensions[1].height = 30
ws.freeze_panes = "B2"

tbl = Table(displayName="Holdings", ref=f"A1:{get_column_letter(len(headers))}{nrows}")
tbl.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=False,
                                    showColumnStripes=False)
ws.add_table(tbl)

# ============================================================
# Sheet 2 — Earnings Schedule (chronological)
# ============================================================
ws2 = wb.create_sheet("Earnings Schedule")
sched = []
for s in rows:
    for e in (live_of(s["t"]).get("earnings") or []):
        d = dt.date.fromisoformat(e["date"])
        sched.append((d, s["t"], s.get("name", ""), s["sector"],
                      e.get("time") or "", "Estimated" if e.get("estimated") else "Confirmed",
                      e.get("source", "")))
sched.sort(key=lambda x: (x[0], x[1]))

sh_headers = ["Date", "Weekday", "Ticker", "Company", "Sector", "Time", "Status", "Source"]
for j, h in enumerate(sh_headers, 1):
    hdr_cell(ws2.cell(1, j), h)
rr = 2
for d, t, name, sec, time, status, source in sched:
    time_txt = {"BMO": "Before open", "AMC": "After close"}.get(time, "")
    vals = [d, d.strftime("%A"), t, name, sec, time_txt, status, source]
    for j, v in enumerate(vals, 1):
        c = ws2.cell(rr, j, value=v)
        c.font = base_font(size=10, bold=(j == 3))
        c.border = BORDER
        if status == "Estimated":
            c.fill = EST_FILL
    ws2.cell(rr, 1).number_format = 'mmm dd, yyyy'
    stc = ws2.cell(rr, 7)
    stc.font = base_font(size=9, italic=(status == "Estimated"),
                         bold=(status == "Confirmed"),
                         color=(GREY if status == "Estimated" else GREEN))
    rr += 1
for j, w in enumerate([14, 12, 9, 24, 22, 13, 11, 14], 1):
    ws2.column_dimensions[get_column_letter(j)].width = w
ws2.row_dimensions[1].height = 24
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:H{rr-1}"

# ============================================================
# Sheet 3 — Calendar (month grids Jul-Dec 2026)
# ============================================================
ws3 = wb.create_sheet("Calendar")
by_day = {}
for d, t, name, sec, time, status, source in sched:
    by_day.setdefault(d, []).append((t, status))

DAYW = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
for j in range(7):
    ws3.column_dimensions[get_column_letter(j + 1)].width = 20

title = ws3.cell(1, 1, "Earnings Calendar · Jul–Dec 2026")
title.font = base_font(bold=True, size=16, color=NAVY)
row_ptr = 3
import calendar as _cal
for month in range(7, 13):
    mname = dt.date(2026, month, 1).strftime("%B %Y")
    mc = ws3.cell(row_ptr, 1, mname)
    mc.font = base_font(bold=True, size=13, color=GOLD)
    row_ptr += 1
    for j, dn in enumerate(DAYW, 1):
        c = ws3.cell(row_ptr, j, dn)
        c.font = base_font(bold=True, color="ffffff", size=10)
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center")
    row_ptr += 1
    weeks = _cal.Calendar(firstweekday=0).monthdatescalendar(2026, month)
    for wk in weeks:
        ws3.row_dimensions[row_ptr].height = 58
        for j, day in enumerate(wk, 1):
            c = ws3.cell(row_ptr, j)
            c.border = BORDER
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if day.month != month:
                c.fill = PatternFill("solid", fgColor="fbfaf7")
                continue
            reporters = by_day.get(day, [])
            head = f"{day.day}"
            if reporters:
                tickers = " ".join(t for t, _ in reporters)
                c.value = f"{day.day}\n{tickers}"
            else:
                c.value = head
            c.font = base_font(size=9, color=INK)
        row_ptr += 1
    row_ptr += 2  # gap between months
ws3.freeze_panes = "A2"

# ============================================================
# Sheet 4 — About
# ============================================================
ws4 = wb.create_sheet("About")
ws4.column_dimensions["A"].width = 100
asof = "unknown"
if REFRESHED:
    asof = dt.datetime.fromisoformat(REFRESHED.replace("Z", "+00:00")).strftime("%B %d, %Y at %H:%M UTC")
lines = [
    ("Earnings Calendar — Data Export", base_font(bold=True, size=16, color=NAVY)),
    (f"Data as of {asof}", base_font(italic=True, size=11, color=GREY)),
    ("", None),
    (f"{len(rows)} equities across {len(set(s['sector'] for s in rows))} GICS sectors.", base_font(size=11)),
    ("", None),
    ("Sheets:", base_font(bold=True, size=12, color=GOLD)),
    ("  • All Holdings — every stock with price, 30/60-day performance, model memberships,", base_font(size=11)),
    ("     analyst ratings & targets (Wells Fargo / Evercore ISI / Morningstar), and 2026 earnings dates.", base_font(size=11)),
    ("  • Earnings Schedule — every report chronologically; filter by sector, status, or date.", base_font(size=11)),
    ("  • Calendar — month-by-month grid, Jul–Dec 2026.", base_font(size=11)),
    ("", None),
    ("Notes:", base_font(bold=True, size=12, color=GOLD)),
    ("  • Prices and confirmed dates refresh automatically each weekday from public market data.", base_font(size=11)),
    ("  • Dates marked \"est\"/\"Estimated\" are seasonal projections until the company confirms.", base_font(size=11)),
    ("  • Model memberships, ratings, and targets were digitized from printed sheets; a ⚠ flags", base_font(size=11)),
    ("     any entry that was hard to read and is worth verifying.", base_font(size=11)),
    ("  • Informational only; not personalized investment advice.", base_font(italic=True, size=10, color=GREY)),
]
for i, (txt, f) in enumerate(lines, 1):
    c = ws4.cell(i, 1, txt)
    if f:
        c.font = f

out_path = os.path.join(ROOT, "earnings-data.xlsx")
wb.save(out_path)
print(f"wrote {out_path}: {nrows} holdings, {len(sched)} scheduled reports")
